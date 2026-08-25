from django.contrib.auth import get_user_model
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import ExamHall, Camera, OfflineExam, StudentZone, HallEnrollment
import io
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
from .serializers import (
    ExamHallSerializer,
    CameraSerializer,
    OfflineExamSerializer,
    StudentZoneSerializer,
    HallEnrollmentSerializer
)

User = get_user_model()


# ─── ExamHall ─────────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def exam_hall_list(request):
    if request.method == 'GET':
        if request.user.is_superuser:
            halls = ExamHall.objects.filter(is_active=True)
        else:
            halls = ExamHall.objects.filter(is_active=True, professor=request.user)
        serializer = ExamHallSerializer(halls, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = ExamHallSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(professor=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def exam_hall_detail(request, pk):
    try:
        if request.user.is_superuser:
            hall = ExamHall.objects.get(pk=pk)
        else:
            hall = ExamHall.objects.get(pk=pk, professor=request.user)
    except ExamHall.DoesNotExist:
        return Response({'error': 'Hall not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = ExamHallSerializer(hall)
        return Response(serializer.data)

    elif request.method == 'PUT':
        serializer = ExamHallSerializer(hall, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        hall.delete()
        return Response({'message': 'Hall deleted'}, status=status.HTTP_200_OK)


# ─── Camera ───────────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def camera_list(request, hall_id):
    try:
        hall = ExamHall.objects.get(pk=hall_id)
    except ExamHall.DoesNotExist:
        return Response({'error': 'Hall not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        cameras = Camera.objects.filter(hall=hall)
        serializer = CameraSerializer(cameras, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = CameraSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(hall=hall)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def camera_detail(request, pk):
    try:
        camera = Camera.objects.get(pk=pk)
    except Camera.DoesNotExist:
        return Response({'error': 'Camera not found'}, status=status.HTTP_404_NOT_FOUND)

    camera.delete()
    return Response({'message': 'Camera deleted'}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def camera_snapshot(request, pk):
    """
    GET /api/hardware/monitoring/cameras/<pk>/snapshot/

    Captures a single frame from the camera's stream_url and returns it as a
    base64-encoded JPEG.

    Response (200):
        {
          "snapshot": "data:image/jpeg;base64,...",
          "width": 1280,
          "height": 720,
          "camera_name": "Front Left",
          "stream_url": "rtsp://..."
        }

    Response (503 / 400):
        { "error": "..." }

    Works for both RTSP URLs and integer webcam index strings (e.g. "0").
    """
    try:
        camera = Camera.objects.get(pk=pk)
    except Camera.DoesNotExist:
        return Response({'error': 'Camera not found'}, status=status.HTTP_404_NOT_FOUND)

    stream_url = camera.stream_url
    if not stream_url and stream_url != 0:
        return Response(
            {'error': 'Camera has no stream URL configured.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        import cv2
        import base64
        import time as _time

        import os
        from django.conf import settings
        from django.http import HttpResponse
        
        frame = None
        ret = False
        
        # Check if the local AI loop is running and has saved a FRESH frame (< 3 seconds old)
        scratch_dir = os.path.join(settings.BASE_DIR, "scratch")
        debug_path = os.path.join(scratch_dir, f"camera_{pk}_latest.jpg")
        
        if os.path.exists(debug_path):
            age = _time.time() - os.path.getmtime(debug_path)
            if age < 3.0:   # only use disk frame if it was written within the last 3 s
                try:
                    frame = cv2.imread(debug_path)
                    if frame is not None:
                        ret = True
                except Exception:
                    pass

        if not ret:
            # Capture a fresh frame directly from the stream
            source_str = str(stream_url).strip()
            is_local = source_str.isdigit()

            if is_local:
                return Response(
                    {'error': 'AI is initializing camera. Please wait...'},
                    status=status.HTTP_503_SERVICE_UNAVAILABLE
                )

            # Fast single-frame grab for IP cameras (grab then retrieve avoids decoding every internal frame)
            cap = cv2.VideoCapture(stream_url)
            if not cap.isOpened():
                return Response(
                    {'error': f'Cannot open camera stream: {stream_url}'},
                    status=status.HTTP_503_SERVICE_UNAVAILABLE
                )

            # Discard buffered frames by grabbing several times, then retrieve the latest one
            for _ in range(5):
                cap.grab()
            ret, frame = cap.retrieve()
            cap.release()

            if not ret or frame is None:
                return Response(
                    {'error': 'Camera opened but returned an empty frame. Stream may be initialising.'},
                    status=status.HTTP_503_SERVICE_UNAVAILABLE
                )

        h, w = frame.shape[:2]

        # Encode to JPEG and then base64
        ok, buf = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
        if not ok:
            return Response({'error': 'Failed to encode frame to JPEG.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        b64 = base64.b64encode(buf.tobytes()).decode('utf-8')
        data_url = f'data:image/jpeg;base64,{b64}'

        response = Response({
            'snapshot':    data_url,
            'width':       w,
            'height':      h,
            'camera_name': camera.name,
            'stream_url':  str(stream_url),
        })
        # Prevent browser and proxy caching so every poll gets a fresh frame
        response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

    except Exception as exc:
        return Response(
            {'error': f'Snapshot failed: {str(exc)}'},
            status=status.HTTP_503_SERVICE_UNAVAILABLE
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def zone_snapshot(request, pk):
    """
    GET /api/hardware/monitoring/zones/<pk>/snapshot/

    Returns a cropped JPEG snapshot for a specific StudentZone.
    Grabs the full camera frame and crops it to the zone's (x1, y1, x2, y2).

    Response (200):
        {
          "snapshot": "data:image/jpeg;base64,...",
          "width": <zone_width>,
          "height": <zone_height>,
          "zone_id": <pk>,
          "student_name": "...",
          "student_code": "..."
        }
    """
    from .models import StudentZone
    try:
        zone = StudentZone.objects.select_related('camera').get(pk=pk)
    except StudentZone.DoesNotExist:
        return Response({'error': 'Zone not found'}, status=status.HTTP_404_NOT_FOUND)

    camera = zone.camera
    if not camera or not camera.stream_url:
        return Response({'error': 'Zone has no camera configured.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        import cv2, base64, time as _time, os
        from django.conf import settings

        frame = None
        ret = False

        # Try disk-cached frame first (written by run_local_ai)
        scratch_dir = os.path.join(settings.BASE_DIR, 'scratch')
        debug_path = os.path.join(scratch_dir, f'camera_{camera.pk}_latest.jpg')
        if os.path.exists(debug_path):
            age = _time.time() - os.path.getmtime(debug_path)
            if age < 3.0:
                frame = cv2.imread(debug_path)
                if frame is not None:
                    ret = True

        if not ret:
            source_str = str(camera.stream_url).strip()
            is_local = source_str.isdigit()
            if is_local:
                return Response({'error': 'AI is initializing camera. Please wait...'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

            cap = cv2.VideoCapture(camera.stream_url)
            if not cap.isOpened():
                return Response({'error': f'Cannot open camera stream'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
            for _ in range(5):
                cap.grab()
            ret, frame = cap.retrieve()
            cap.release()

            if not ret or frame is None:
                return Response({'error': 'Camera returned an empty frame.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        # Crop to zone coordinates
        h, w = frame.shape[:2]
        x1 = max(0, min(int(zone.x1), w - 1))
        y1 = max(0, min(int(zone.y1), h - 1))
        x2 = max(x1 + 1, min(int(zone.x2), w))
        y2 = max(y1 + 1, min(int(zone.y2), h))
        crop = frame[y1:y2, x1:x2]

        if crop.size == 0:
            return Response({'error': 'Zone coordinates produce an empty crop.'}, status=status.HTTP_400_BAD_REQUEST)

        ok, buf = cv2.imencode('.jpg', crop, [cv2.IMWRITE_JPEG_QUALITY, 82])
        if not ok:
            return Response({'error': 'Failed to encode crop to JPEG.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        b64 = base64.b64encode(buf.tobytes()).decode('utf-8')
        data_url = f'data:image/jpeg;base64,{b64}'

        ch, cw = crop.shape[:2]
        response = Response({
            'snapshot':     data_url,
            'width':        cw,
            'height':       ch,
            'zone_id':      zone.pk,
            'student_name': zone.student_name or '',
            'student_code': zone.student_code or '',
        })
        response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

    except Exception as exc:
        return Response({'error': f'Zone snapshot failed: {str(exc)}'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)


# ─── OfflineExam ──────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def offline_exam_list(request):
    if request.method == 'GET':
        if request.user.is_superuser:
            exams = OfflineExam.objects.all()
        else:
            exams = OfflineExam.objects.filter(professor=request.user)
        serializer = OfflineExamSerializer(exams, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = OfflineExamSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(professor=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def offline_exam_detail(request, pk):
    try:
        if request.user.is_superuser:
            exam = OfflineExam.objects.get(pk=pk)
        else:
            exam = OfflineExam.objects.get(pk=pk, professor=request.user)
    except OfflineExam.DoesNotExist:
        return Response({'error': 'Exam not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = OfflineExamSerializer(exam)
        return Response(serializer.data)

    elif request.method == 'PUT':
        serializer = OfflineExamSerializer(exam, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        exam.delete()
        return Response({'message': 'Exam deleted'}, status=status.HTTP_200_OK)


# ─── StudentZone ──────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def student_zone_list(request, exam_id):
    # Resolve the hall from the offline exam so the URL stays exam-based
    try:
        exam = OfflineExam.objects.get(pk=exam_id)
    except OfflineExam.DoesNotExist:
        return Response({'error': 'Exam not found'}, status=status.HTTP_404_NOT_FOUND)

    hall = exam.hall

    if request.method == 'GET':
        zones = StudentZone.objects.filter(hall=hall)
        camera_id = request.query_params.get('camera')
        if camera_id:
            zones = zones.filter(camera=camera_id)
        serializer = StudentZoneSerializer(zones, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        data = request.data.copy()

        # If a student (enrollment) id was sent, pull name/code from HallEnrollment
        enrollment_id = data.get('student')
        if enrollment_id:
            try:
                enrollment = HallEnrollment.objects.get(pk=enrollment_id, hall=hall)
                data['student_name'] = enrollment.student_name
                data['student_code'] = enrollment.student_code
            except HallEnrollment.DoesNotExist:
                pass  # leave name/code as sent by client

        camera_id = data.get('camera')
        if camera_id and not Camera.objects.filter(pk=camera_id, hall=hall).exists():
            return Response(
                {'error': 'Camera does not belong to this hall'},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = StudentZoneSerializer(data=data)
        if serializer.is_valid():
            serializer.save(hall=hall)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def student_zone_detail(request, pk):
    try:
        zone = StudentZone.objects.get(pk=pk)
    except StudentZone.DoesNotExist:
        return Response({'error': 'Zone not found'}, status=status.HTTP_404_NOT_FOUND)

    zone.delete()
    return Response({'message': 'Zone deleted'}, status=status.HTTP_200_OK)


# ─── HallEnrollment ───────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def hall_enrollment_list(request, hall_id):
    try:
        hall = ExamHall.objects.get(pk=hall_id)
    except ExamHall.DoesNotExist:
        return Response({'error': 'Hall not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        enrollments = HallEnrollment.objects.filter(hall=hall)
        serializer  = HallEnrollmentSerializer(enrollments, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = HallEnrollmentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(hall=hall)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE', 'PUT'])
@permission_classes([IsAuthenticated])
def hall_enrollment_detail(request, pk):
    try:
        enrollment = HallEnrollment.objects.get(pk=pk)
    except HallEnrollment.DoesNotExist:
        return Response({'error': 'Enrollment not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'PUT':
        serializer = HallEnrollmentSerializer(enrollment, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    enrollment.delete()
    return Response({'message': 'Student removed from hall'}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_list_view(request):
    students = User.objects.filter(role='student').values('id', 'first_name', 'last_name', 'email')
    data = [
        {
            'id': s['id'],
            'name': f"{s['first_name']} {s['last_name']}".strip() or s['email'],
            'email': s['email'],
        }
        for s in students
    ]
    return Response(data)


# ─── Bulk Enroll from Excel ───────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_enroll_from_excel(request, hall_id):
    """
    Accept a multipart/form-data POST with a single file field named 'file'.
    The Excel sheet must have (at minimum) these columns in any order:
        Student Name | ID | Seat Number
    Returns a summary: { created, skipped, errors }.
    """
    if not OPENPYXL_AVAILABLE:
        return Response(
            {'error': 'openpyxl is not installed on the server. Run: pip install openpyxl'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

    try:
        hall = ExamHall.objects.get(pk=hall_id)
    except ExamHall.DoesNotExist:
        return Response({'error': 'Hall not found'}, status=status.HTTP_404_NOT_FOUND)

    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return Response({'error': 'No file uploaded. Send the Excel as a multipart field named "file".'}, status=status.HTTP_400_BAD_REQUEST)

    # Accept .xlsx and .xls
    filename = uploaded_file.name.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        return Response({'error': 'Only .xlsx or .xls files are accepted.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        file_bytes = uploaded_file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
    except Exception as exc:
        return Response({'error': f'Could not read Excel file: {exc}'}, status=status.HTTP_400_BAD_REQUEST)

    # Normalise header row — expect: Student Name, ID, Seat Number
    header_row = [str(cell.value).strip().lower() if cell.value is not None else '' for cell in ws[1]]

    ALIASES = {
        'student name': ['student name', 'name', 'student_name', 'studentname'],
        'id':           ['id', 'student id', 'student_id', 'studentid', 'code', 'student code'],
        'seat number':  ['seat number', 'seat_number', 'seat no', 'seat', 'seatnumber', 'seat no.'],
    }

    col_index = {}
    for canonical, aliases in ALIASES.items():
        for i, h in enumerate(header_row):
            if h in aliases:
                col_index[canonical] = i
                break

    missing = [k for k in ['student name', 'id'] if k not in col_index]
    if missing:
        return Response(
            {'error': f'Missing required column(s): {", ".join(missing)}. Found headers: {", ".join(header_row)}'},
            status=status.HTTP_400_BAD_REQUEST
        )

    created_count = 0
    skipped_count = 0
    row_errors = []
    
    seat_counter = 1

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        student_name = str(row[col_index['student name']] or '').strip()
        student_code = str(row[col_index['id']] or '').strip()
        seat_number  = str(row[col_index.get('seat number', -1)] or '').strip() if 'seat number' in col_index else ''

        if not student_name and not student_code:
            continue  # blank row

        if not student_code:
            row_errors.append({'row': row_num, 'reason': f'Missing student ID for "{student_name}"'})
            continue
            
        if not seat_number:
            seat_number = str(seat_counter)
            
        seat_counter += 1

        _, created = HallEnrollment.objects.get_or_create(
            hall=hall,
            student_code=student_code,
            defaults={
                'student_name': student_name,
                'seat_number':  seat_number or None,
            }
        )
        if created:
            created_count += 1
        else:
            skipped_count += 1

    return Response({
        'created': created_count,
        'skipped': skipped_count,
        'errors':  row_errors,
        'message': f'{created_count} student(s) enrolled, {skipped_count} already existed.',
    }, status=status.HTTP_200_OK)