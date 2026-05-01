from django.contrib.auth import get_user_model
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import ExamHall, Camera, OfflineExam, StudentZone, HallEnrollment
import io
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
from .serializers import (
    ExamHallSerializer,
    CameraSerializer,
    OfflineExamSerializer,
    StudentZoneSerializer,
    HallEnrollmentSerializer
)

User = get_user_model()


# ─── ExamHall ─────────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def exam_hall_list(request):
    if request.method == 'GET':
        if request.user.is_superuser:
            halls = ExamHall.objects.filter(is_active=True)
        else:
            halls = ExamHall.objects.filter(is_active=True, professor=request.user)
        serializer = ExamHallSerializer(halls, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = ExamHallSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(professor=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def exam_hall_detail(request, pk):
    try:
        if request.user.is_superuser:
            hall = ExamHall.objects.get(pk=pk)
        else:
            hall = ExamHall.objects.get(pk=pk, professor=request.user)
    except ExamHall.DoesNotExist:
        return Response({'error': 'Hall not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = ExamHallSerializer(hall)
        return Response(serializer.data)

    elif request.method == 'PUT':
        serializer = ExamHallSerializer(hall, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        hall.delete()
        return Response({'message': 'Hall deleted'}, status=status.HTTP_200_OK)


# ─── Camera ───────────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def camera_list(request, hall_id):
    try:
        hall = ExamHall.objects.get(pk=hall_id)
    except ExamHall.DoesNotExist:
        return Response({'error': 'Hall not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        cameras = Camera.objects.filter(hall=hall)
        serializer = CameraSerializer(cameras, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = CameraSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(hall=hall)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def camera_detail(request, pk):
    try:
        camera = Camera.objects.get(pk=pk)
    except Camera.DoesNotExist:
        return Response({'error': 'Camera not found'}, status=status.HTTP_404_NOT_FOUND)

    camera.delete()
    return Response({'message': 'Camera deleted'}, status=status.HTTP_200_OK)


# ─── OfflineExam ──────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def offline_exam_list(request):
    if request.method == 'GET':
        if request.user.is_superuser:
            exams = OfflineExam.objects.all()
        else:
            exams = OfflineExam.objects.filter(professor=request.user)
        serializer = OfflineExamSerializer(exams, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = OfflineExamSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(professor=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def offline_exam_detail(request, pk):
    try:
        if request.user.is_superuser:
            exam = OfflineExam.objects.get(pk=pk)
        else:
            exam = OfflineExam.objects.get(pk=pk, professor=request.user)
    except OfflineExam.DoesNotExist:
        return Response({'error': 'Exam not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = OfflineExamSerializer(exam)
        return Response(serializer.data)

    elif request.method == 'PUT':
        serializer = OfflineExamSerializer(exam, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        exam.delete()
        return Response({'message': 'Exam deleted'}, status=status.HTTP_200_OK)


# ─── StudentZone ──────────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def student_zone_list(request, exam_id):
    # Resolve the hall from the offline exam so the URL stays exam-based
    try:
        exam = OfflineExam.objects.get(pk=exam_id)
    except OfflineExam.DoesNotExist:
        return Response({'error': 'Exam not found'}, status=status.HTTP_404_NOT_FOUND)

    hall = exam.hall

    if request.method == 'GET':
        zones = StudentZone.objects.filter(hall=hall)
        camera_id = request.query_params.get('camera')
        if camera_id:
            zones = zones.filter(camera=camera_id)
        serializer = StudentZoneSerializer(zones, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        data = request.data.copy()

        # If a student (enrollment) id was sent, pull name/code from HallEnrollment
        enrollment_id = data.get('student')
        if enrollment_id:
            try:
                enrollment = HallEnrollment.objects.get(pk=enrollment_id, hall=hall)
                data['student_name'] = enrollment.student_name
                data['student_code'] = enrollment.student_code
            except HallEnrollment.DoesNotExist:
                pass  # leave name/code as sent by client

        camera_id = data.get('camera')
        if camera_id and not Camera.objects.filter(pk=camera_id, hall=hall).exists():
            return Response(
                {'error': 'Camera does not belong to this hall'},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = StudentZoneSerializer(data=data)
        if serializer.is_valid():
            serializer.save(hall=hall)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def student_zone_detail(request, pk):
    try:
        zone = StudentZone.objects.get(pk=pk)
    except StudentZone.DoesNotExist:
        return Response({'error': 'Zone not found'}, status=status.HTTP_404_NOT_FOUND)

    zone.delete()
    return Response({'message': 'Zone deleted'}, status=status.HTTP_200_OK)


# ─── HallEnrollment ───────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def hall_enrollment_list(request, hall_id):
    try:
        hall = ExamHall.objects.get(pk=hall_id)
    except ExamHall.DoesNotExist:
        return Response({'error': 'Hall not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        enrollments = HallEnrollment.objects.filter(hall=hall)
        serializer  = HallEnrollmentSerializer(enrollments, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = HallEnrollmentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(hall=hall)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE', 'PUT'])
@permission_classes([IsAuthenticated])
def hall_enrollment_detail(request, pk):
    try:
        enrollment = HallEnrollment.objects.get(pk=pk)
    except HallEnrollment.DoesNotExist:
        return Response({'error': 'Enrollment not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'PUT':
        serializer = HallEnrollmentSerializer(enrollment, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    enrollment.delete()
    return Response({'message': 'Student removed from hall'}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_list_view(request):
    students = User.objects.filter(role='student').values('id', 'first_name', 'last_name', 'email')
    data = [
        {
            'id': s['id'],
            'name': f"{s['first_name']} {s['last_name']}".strip() or s['email'],
            'email': s['email'],
        }
        for s in students
    ]
    return Response(data)


# ─── Bulk Enroll from Excel ───────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_enroll_from_excel(request, hall_id):
    """
    Accept a multipart/form-data POST with a single file field named 'file'.
    The Excel sheet must have (at minimum) these columns in any order:
        Student Name | ID | Seat Number
    Returns a summary: { created, skipped, errors }.
    """
    if not OPENPYXL_AVAILABLE:
        return Response(
            {'error': 'openpyxl is not installed on the server. Run: pip install openpyxl'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

    try:
        hall = ExamHall.objects.get(pk=hall_id)
    except ExamHall.DoesNotExist:
        return Response({'error': 'Hall not found'}, status=status.HTTP_404_NOT_FOUND)

    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return Response({'error': 'No file uploaded. Send the Excel as a multipart field named "file".'}, status=status.HTTP_400_BAD_REQUEST)

    # Accept .xlsx and .xls
    filename = uploaded_file.name.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        return Response({'error': 'Only .xlsx or .xls files are accepted.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        file_bytes = uploaded_file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
    except Exception as exc:
        return Response({'error': f'Could not read Excel file: {exc}'}, status=status.HTTP_400_BAD_REQUEST)

    # Normalise header row — expect: Student Name, ID, Seat Number
    header_row = [str(cell.value).strip().lower() if cell.value is not None else '' for cell in ws[1]]

    ALIASES = {
        'student name': ['student name', 'name', 'student_name', 'studentname'],
        'id':           ['id', 'student id', 'student_id', 'studentid', 'code', 'student code'],
        'seat number':  ['seat number', 'seat_number', 'seat no', 'seat', 'seatnumber', 'seat no.'],
    }

    col_index = {}
    for canonical, aliases in ALIASES.items():
        for i, h in enumerate(header_row):
            if h in aliases:
                col_index[canonical] = i
                break

    missing = [k for k in ['student name', 'id'] if k not in col_index]
    if missing:
        return Response(
            {'error': f'Missing required column(s): {", ".join(missing)}. Found headers: {", ".join(header_row)}'},
            status=status.HTTP_400_BAD_REQUEST
        )

    created_count = 0
    skipped_count = 0
    row_errors = []
    
    seat_counter = 1

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        student_name = str(row[col_index['student name']] or '').strip()
        student_code = str(row[col_index['id']] or '').strip()
        seat_number  = str(row[col_index.get('seat number', -1)] or '').strip() if 'seat number' in col_index else ''

        if not student_name and not student_code:
            continue  # blank row

        if not student_code:
            row_errors.append({'row': row_num, 'reason': f'Missing student ID for "{student_name}"'})
            continue
            
        if not seat_number:
            seat_number = str(seat_counter)
            
        seat_counter += 1

        _, created = HallEnrollment.objects.get_or_create(
            hall=hall,
            student_code=student_code,
            defaults={
                'student_name': student_name,
                'seat_number':  seat_number or None,
            }
        )
        if created:
            created_count += 1
        else:
            skipped_count += 1

    return Response({
        'created': created_count,
        'skipped': skipped_count,
        'errors':  row_errors,
        'message': f'{created_count} student(s) enrolled, {skipped_count} already existed.',
    }, status=status.HTTP_200_OK)